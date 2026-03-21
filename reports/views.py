from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Count, Avg, Max, Min, F, Q
from django.utils import timezone
from decimal import Decimal
from shops.models import Shop
from pos.models import Sale, SaleItem
from expenses.models import Expense
from stock.models import StockLevel, StockMovement
from customers.models import Customer, CreditPayment
from products.models import Product
import datetime, json


PERIODS = [
    ('today', 'Today'), ('yesterday', 'Yesterday'),
    ('week', 'This week'), ('last_week', 'Last week'),
    ('month', 'This month'), ('last_month', 'Last month'),
    ('quarter', 'Quarter'), ('year', 'Year'),
]

def base_ctx(report, period, start, end):
    return {'report': report, 'period': period, 'start': start, 'end': end, 'periods': PERIODS}

def get_current_shop(request):
    shop_id = request.session.get('current_shop_id')
    return Shop.objects.filter(id=shop_id).first() if shop_id else None


def get_date_range(request):
    period = request.GET.get('period', 'month')
    today  = timezone.now().date()
    if period == 'today':
        return today, today, period
    elif period == 'yesterday':
        y = today - datetime.timedelta(days=1)
        return y, y, period
    elif period == 'week':
        start = today - datetime.timedelta(days=today.weekday())
        return start, today, period
    elif period == 'last_week':
        end   = today - datetime.timedelta(days=today.weekday() + 1)
        start = end - datetime.timedelta(days=6)
        return start, end, period
    elif period == 'month':
        return today.replace(day=1), today, period
    elif period == 'last_month':
        end   = today.replace(day=1) - datetime.timedelta(days=1)
        start = end.replace(day=1)
        return start, end, period
    elif period == 'quarter':
        q     = (today.month - 1) // 3
        start = today.replace(month=q * 3 + 1, day=1)
        return start, today, period
    elif period == 'year':
        return today.replace(month=1, day=1), today, period
    elif period == 'custom':
        try:
            start = datetime.date.fromisoformat(request.GET.get('start', str(today)))
            end   = datetime.date.fromisoformat(request.GET.get('end', str(today)))
            return start, end, period
        except Exception:
            return today, today, period
    return today.replace(day=1), today, period


def pct(a, b):
    """% change from b to a."""
    try:
        if float(b) == 0:
            return None
        return round(((float(a) - float(b)) / float(b)) * 100, 1)
    except Exception:
        return None


def day_series(shop, start, end, model_qs):
    """Build a list of {date, value} dicts for every day in range."""
    series = []
    d = start
    while d <= end:
        val = model_qs.filter(created_at__date=d).aggregate(t=Sum('total'))['t'] or 0
        series.append({'date': d.strftime('%d %b'), 'value': float(val)})
        d += datetime.timedelta(days=1)
    return series


# ── REPORT ROUTER ────────────────────────────────────────────────
@login_required
def reports_dashboard(request):
    shop = get_current_shop(request)
    if not shop:
        return redirect('shop_select')
    report = request.GET.get('report', 'overview')
    view_map = {
        'overview':    report_overview,
        'sales':       report_sales,
        'products':    report_products,
        'expenses':    report_expenses,
        'customers':   report_customers,
        'stock':       report_stock,
        'profit':      report_profit,
        'forecast':    report_forecast,
    }
    fn = view_map.get(report, report_overview)
    return fn(request, shop)


# ── 1. OVERVIEW ──────────────────────────────────────────────────
def report_overview(request, shop):
    start, end, period = get_date_range(request)
    today = timezone.now().date()

    sales_qs   = Sale.objects.filter(shop=shop, created_at__date__range=(start, end), status='completed')
    items_qs   = SaleItem.objects.filter(sale__in=sales_qs)
    exp_qs     = Expense.objects.filter(shop=shop, date__range=(start, end))

    revenue    = sales_qs.aggregate(t=Sum('total'))['t'] or Decimal('0')
    tax        = sales_qs.aggregate(t=Sum('tax_amount'))['t'] or Decimal('0')
    txn_count  = sales_qs.count()
    avg_sale   = revenue / txn_count if txn_count else Decimal('0')
    cogs       = sum((i.buying_price * i.quantity for i in items_qs), Decimal('0'))
    gross      = revenue - cogs
    expenses   = exp_qs.aggregate(t=Sum('amount'))['t'] or Decimal('0')
    net        = gross - expenses
    margin     = round((float(gross) / float(revenue) * 100), 1) if revenue else 0

    # Payment split
    cash   = sales_qs.filter(payment_method='cash').aggregate(t=Sum('total'))['t'] or Decimal('0')
    mpesa  = sales_qs.filter(payment_method='mpesa').aggregate(t=Sum('total'))['t'] or Decimal('0')
    credit = sales_qs.filter(payment_method='credit').aggregate(t=Sum('total'))['t'] or Decimal('0')

    # Daily revenue series for chart
    series = []
    d = start
    while d <= end:
        v = Sale.objects.filter(shop=shop, created_at__date=d, status='completed').aggregate(t=Sum('total'))['t'] or 0
        series.append({'date': d.strftime('%d %b'), 'revenue': float(v)})
        d += datetime.timedelta(days=1)

    # Best & worst day
    best_day  = max(series, key=lambda x: x['revenue'], default=None)
    worst_day = min((s for s in series if s['revenue'] > 0), key=lambda x: x['revenue'], default=None)

    # Expense breakdown
    exp_cats = exp_qs.values('category__name').annotate(t=Sum('amount')).order_by('-t')

    payment_rows = [('Cash', cash, '#6146c1'), ('M-Pesa', mpesa, '#0D512B'), ('Credit', credit, '#EB5993')]
    ctx = dict(
        shop=shop, report='overview', period=period, start=start, end=end,
        periods=PERIODS,
        payment_rows=payment_rows,
        revenue=revenue, tax=tax, txn_count=txn_count, avg_sale=avg_sale,
        cogs=cogs, gross=gross, expenses=expenses, net=net, margin=margin,
        cash=cash, mpesa=mpesa, credit=credit,
        series=json.dumps(series), best_day=best_day, worst_day=worst_day,
        exp_cats=exp_cats,
    )
    return render(request, 'reports/overview.html', ctx)


# ── 2. SALES DEEP DIVE ───────────────────────────────────────────
def report_sales(request, shop):
    start, end, period = get_date_range(request)

    sales_qs  = Sale.objects.filter(shop=shop, created_at__date__range=(start, end), status='completed')
    revenue   = sales_qs.aggregate(t=Sum('total'))['t'] or Decimal('0')
    txn_count = sales_qs.count()
    avg_sale  = revenue / txn_count if txn_count else Decimal('0')
    max_sale  = sales_qs.aggregate(m=Max('total'))['m'] or Decimal('0')
    min_sale  = sales_qs.aggregate(m=Min('total'))['m'] or Decimal('0')

    # Hourly distribution
    hourly = [0] * 24
    for s in sales_qs.only('created_at'):
        hourly[s.created_at.hour] += 1

    # Day-of-week distribution
    dow_labels = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    dow = [0] * 7
    dow_rev = [0.0] * 7
    for s in sales_qs.only('created_at', 'total'):
        dow[s.created_at.weekday()] += 1
        dow_rev[s.created_at.weekday()] += float(s.total)

    # Daily series
    series = []
    d = start
    while d <= end:
        qs = Sale.objects.filter(shop=shop, created_at__date=d, status='completed')
        v  = qs.aggregate(t=Sum('total'))['t'] or 0
        c  = qs.count()
        series.append({'date': d.strftime('%d %b'), 'revenue': float(v), 'count': c})
        d += datetime.timedelta(days=1)

    # Payment methods
    pay_breakdown = sales_qs.values('payment_method').annotate(
        count=Count('id'), total=Sum('total')
    ).order_by('-total')

    # Recent transactions
    recent = sales_qs.select_related('customer', 'cashier').order_by('-created_at')[:20]

    ctx = dict(
        shop=shop, report='sales', period=period, start=start, end=end,
        periods=PERIODS,
        revenue=revenue, txn_count=txn_count, avg_sale=avg_sale,
        max_sale=max_sale, min_sale=min_sale,
        hourly=json.dumps(hourly),
        dow_labels=json.dumps(dow_labels), dow=json.dumps(dow), dow_rev=json.dumps(dow_rev),
        series=json.dumps(series),
        pay_breakdown=pay_breakdown,
        recent=recent,
    )
    return render(request, 'reports/sales.html', ctx)


# ── 3. PRODUCT PERFORMANCE ───────────────────────────────────────
def report_products(request, shop):
    start, end, period = get_date_range(request)

    items_qs = SaleItem.objects.filter(
        sale__shop=shop,
        sale__created_at__date__range=(start, end),
        sale__status='completed'
    )

    # Top by revenue
    by_revenue = items_qs.values(
        'product__name', 'product__uom__short_name', 'product__category__name'
    ).annotate(
        qty=Sum('quantity'),
        revenue=Sum('line_total'),
        cogs=Sum(F('buying_price') * F('quantity')),
        txns=Count('sale', distinct=True),
    ).order_by('-revenue')[:20]

    # Attach profit and margin
    top_products = []
    for p in by_revenue:
        profit = float(p['revenue'] or 0) - float(p['cogs'] or 0)
        margin = round(profit / float(p['revenue']) * 100, 1) if p['revenue'] else 0
        top_products.append({**p, 'profit': profit, 'margin': margin})

    # Never sold products this period
    sold_ids = items_qs.values_list('product_id', flat=True).distinct()
    never_sold = Product.objects.filter(shop=shop, is_active=True).exclude(id__in=sold_ids)

    # Category performance
    by_category = items_qs.values('product__category__name').annotate(
        revenue=Sum('line_total'),
        qty=Sum('quantity'),
    ).order_by('-revenue')

    ctx = dict(
        shop=shop, report='products', period=period, start=start, end=end,
        periods=PERIODS,
        top_products=top_products,
        never_sold=never_sold,
        by_category=by_category,
        by_category_json=json.dumps([
            {'cat': p['product__category__name'] or 'Uncategorised', 'rev': float(p['revenue'] or 0)}
            for p in by_category
        ]),
    )
    return render(request, 'reports/products.html', ctx)


# ── 4. EXPENSE ANALYSIS ──────────────────────────────────────────
def report_expenses(request, shop):
    start, end, period = get_date_range(request)

    exp_qs = Expense.objects.filter(shop=shop, date__range=(start, end))
    total  = exp_qs.aggregate(t=Sum('amount'))['t'] or Decimal('0')
    count  = exp_qs.count()
    avg    = total / count if count else Decimal('0')
    max_e  = exp_qs.aggregate(m=Max('amount'))['m'] or Decimal('0')

    by_cat = exp_qs.values('category__name').annotate(
        total=Sum('amount'), count=Count('id')
    ).order_by('-total')

    # Daily series
    series = []
    d = start
    while d <= end:
        v = Expense.objects.filter(shop=shop, date=d).aggregate(t=Sum('amount'))['t'] or 0
        series.append({'date': d.strftime('%d %b'), 'amount': float(v)})
        d += datetime.timedelta(days=1)

    # Revenue vs expense comparison
    rev_qs = Sale.objects.filter(shop=shop, created_at__date__range=(start, end), status='completed')
    revenue = rev_qs.aggregate(t=Sum('total'))['t'] or Decimal('0')
    exp_ratio = round(float(total) / float(revenue) * 100, 1) if revenue else 0

    recent_expenses = exp_qs.select_related('category').order_by('-date', '-created_at')[:20]

    ctx = dict(
        shop=shop, report='expenses', period=period, start=start, end=end,
        periods=PERIODS,
        total=total, count=count, avg=avg, max_e=max_e,
        by_cat=by_cat, series=json.dumps(series),
        revenue=revenue, exp_ratio=exp_ratio,
        recent_expenses=recent_expenses,
        by_cat_json=json.dumps([
            {'cat': c['category__name'] or 'Uncategorised', 'amount': float(c['total'] or 0)}
            for c in by_cat
        ]),
    )
    return render(request, 'reports/expenses.html', ctx)


# ── 5. CUSTOMER INSIGHTS ─────────────────────────────────────────
def report_customers(request, shop):
    start, end, period = get_date_range(request)

    sales_qs = Sale.objects.filter(shop=shop, created_at__date__range=(start, end), status='completed')

    # Top customers by spend
    top_customers = sales_qs.filter(customer__isnull=False).values(
        'customer__name', 'customer__phone'
    ).annotate(
        total_spend=Sum('total'),
        visit_count=Count('id'),
        avg_basket=Avg('total'),
    ).order_by('-total_spend')[:15]

    # Walk-in vs named customer split
    named_count   = sales_qs.filter(customer__isnull=False).count()
    walkin_count  = sales_qs.filter(customer__isnull=True).count()
    named_rev     = sales_qs.filter(customer__isnull=False).aggregate(t=Sum('total'))['t'] or Decimal('0')
    walkin_rev    = sales_qs.filter(customer__isnull=True).aggregate(t=Sum('total'))['t'] or Decimal('0')

    # Outstanding credit
    credit_customers = Customer.objects.filter(shop=shop, credit_balance__gt=0).order_by('-credit_balance')
    total_credit_owed = credit_customers.aggregate(t=Sum('credit_balance'))['t'] or Decimal('0')

    # Credit payments in period
    credit_paid = CreditPayment.objects.filter(
        customer__shop=shop, created_at__date__range=(start, end)
    ).aggregate(t=Sum('amount'))['t'] or Decimal('0')

    # New customers in period
    new_customers = Customer.objects.filter(shop=shop, created_at__date__range=(start, end)).count()

    ctx = dict(
        shop=shop, report='customers', period=period, start=start, end=end,
        periods=PERIODS,
        top_customers=top_customers,
        named_count=named_count, walkin_count=walkin_count,
        named_rev=named_rev, walkin_rev=walkin_rev,
        credit_customers=credit_customers, total_credit_owed=total_credit_owed,
        credit_paid=credit_paid, new_customers=new_customers,
    )
    return render(request, 'reports/customers.html', ctx)


# ── 6. STOCK REPORT ──────────────────────────────────────────────
def report_stock(request, shop):
    start, end, period = get_date_range(request)

    levels = StockLevel.objects.filter(shop=shop).select_related(
        'product', 'product__category', 'product__uom', 'variant'
    )

    # Inventory value
    total_cost_value   = sum(l.quantity * l.product.buying_price for l in levels if l.quantity > 0)
    total_retail_value = sum(l.quantity * l.product.selling_price for l in levels if l.quantity > 0)
    potential_profit   = total_retail_value - total_cost_value

    out_of_stock  = [l for l in levels if l.quantity <= 0]
    low_stock     = [l for l in levels if 0 < l.quantity <= l.product.low_stock_threshold]
    healthy_stock = [l for l in levels if l.quantity > l.product.low_stock_threshold]

    # Stock movements in period
    movements = StockMovement.objects.filter(
        shop=shop, created_at__date__range=(start, end)
    ).select_related('product', 'product__uom').order_by('-created_at')[:30]

    movement_summary = StockMovement.objects.filter(
        shop=shop, created_at__date__range=(start, end)
    ).values('movement_type').annotate(count=Count('id'), qty=Sum('quantity'))

    ctx = dict(
        shop=shop, report='stock', period=period, start=start, end=end,
        periods=PERIODS,
        levels=levels,
        total_cost_value=total_cost_value,
        total_retail_value=total_retail_value,
        potential_profit=potential_profit,
        out_of_stock=out_of_stock,
        low_stock=low_stock,
        healthy_stock=healthy_stock,
        movements=movements,
        movement_summary=movement_summary,
    )
    return render(request, 'reports/stock.html', ctx)


# ── 7. PROFIT & LOSS ─────────────────────────────────────────────
def report_profit(request, shop):
    start, end, period = get_date_range(request)

    sales_qs = Sale.objects.filter(shop=shop, created_at__date__range=(start, end), status='completed')
    items_qs = SaleItem.objects.filter(sale__in=sales_qs)
    exp_qs   = Expense.objects.filter(shop=shop, date__range=(start, end))

    revenue  = sales_qs.aggregate(t=Sum('total'))['t'] or Decimal('0')
    tax      = sales_qs.aggregate(t=Sum('tax_amount'))['t'] or Decimal('0')
    cogs     = sum((i.buying_price * i.quantity for i in items_qs), Decimal('0'))
    gross    = revenue - cogs
    expenses = exp_qs.aggregate(t=Sum('amount'))['t'] or Decimal('0')
    net      = gross - expenses
    margin   = round(float(gross) / float(revenue) * 100, 1) if revenue else 0

    # Expense breakdown for P&L
    exp_cats = exp_qs.values('category__name').annotate(total=Sum('amount')).order_by('-total')

    # Monthly P&L series (last 12 months)
    today = timezone.now().date()
    monthly = []
    for i in range(11, -1, -1):
        month_date  = (today.replace(day=1) - datetime.timedelta(days=1)).replace(day=1) if i > 0 else today.replace(day=1)
        for j in range(i):
            month_date = (month_date.replace(day=1) - datetime.timedelta(days=1)).replace(day=1)
        m_end   = (month_date.replace(day=28) + datetime.timedelta(days=4)).replace(day=1) - datetime.timedelta(days=1)
        m_rev   = Sale.objects.filter(shop=shop, created_at__date__range=(month_date, m_end), status='completed').aggregate(t=Sum('total'))['t'] or 0
        m_exp   = Expense.objects.filter(shop=shop, date__range=(month_date, m_end)).aggregate(t=Sum('amount'))['t'] or 0
        m_items = SaleItem.objects.filter(sale__shop=shop, sale__created_at__date__range=(month_date, m_end), sale__status='completed')
        m_cogs  = sum((i.buying_price * i.quantity for i in m_items), Decimal('0'))
        m_net   = float(m_rev) - float(m_cogs) - float(m_exp)
        monthly.append({'month': month_date.strftime('%b %Y'), 'revenue': float(m_rev), 'net': m_net})

    ctx = dict(
        shop=shop, report='profit', period=period, start=start, end=end,
        periods=PERIODS,
        revenue=revenue, tax=tax, cogs=cogs, gross=gross,
        expenses=expenses, net=net, margin=margin,
        exp_cats=exp_cats,
        monthly_series=json.dumps(monthly),
    )
    return render(request, 'reports/profit.html', ctx)


# ── 8. FORECAST ──────────────────────────────────────────────────
def report_forecast(request, shop):
    today = timezone.now().date()

    # Build 90-day history for trend calculation
    history = []
    for i in range(89, -1, -1):
        d = today - datetime.timedelta(days=i)
        v = Sale.objects.filter(shop=shop, created_at__date=d, status='completed').aggregate(t=Sum('total'))['t'] or 0
        history.append(float(v))

    # 7-day moving average
    def moving_avg(data, window=7):
        result = []
        for i in range(len(data)):
            subset = data[max(0, i - window + 1):i + 1]
            result.append(sum(subset) / len(subset))
        return result

    ma7  = moving_avg(history, 7)
    ma30 = moving_avg(history, 30)

    # Simple linear regression for 30-day forecast
    n = len(history)
    x_mean = (n - 1) / 2
    y_mean = sum(history) / n
    num = sum((i - x_mean) * (history[i] - y_mean) for i in range(n))
    den = sum((i - x_mean) ** 2 for i in range(n))
    slope = num / den if den != 0 else 0
    intercept = y_mean - slope * x_mean

    forecast = []
    for i in range(1, 31):
        predicted = max(0, intercept + slope * (n - 1 + i))
        forecast.append(round(predicted))

    # Labels
    hist_labels = [(today - datetime.timedelta(days=89 - i)).strftime('%d %b') for i in range(90)]
    fore_labels = [(today + datetime.timedelta(days=i)).strftime('%d %b') for i in range(1, 31)]

    # Summary stats
    last7_avg  = sum(history[-7:]) / 7
    last30_avg = sum(history[-30:]) / 30 if len(history) >= 30 else sum(history) / len(history)
    fore30_avg = sum(forecast) / 30
    fore30_total = sum(forecast)
    trend = 'up' if slope > 0 else 'down' if slope < 0 else 'flat'

    # Best performing days of week
    dow_totals = [0.0] * 7
    dow_counts = [0] * 7
    for i, v in enumerate(history):
        d = today - datetime.timedelta(days=89 - i)
        dow_totals[d.weekday()] += v
        dow_counts[d.weekday()] += 1
    dow_avgs = [round(dow_totals[i] / max(dow_counts[i], 1)) for i in range(7)]
    best_dow = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'][dow_avgs.index(max(dow_avgs))]

    ctx = dict(
        shop=shop, report='forecast',
        period='', start='', end='', periods=PERIODS,
        hist_labels=json.dumps(hist_labels),
        history=json.dumps([round(v) for v in history]),
        ma7=json.dumps([round(v) for v in ma7]),
        fore_labels=json.dumps(fore_labels),
        forecast=json.dumps(forecast),
        trend=trend, slope=round(slope),
        last7_avg=round(last7_avg),
        last30_avg=round(last30_avg),
        fore30_avg=round(fore30_avg),
        fore30_total=fore30_total,
        best_dow=best_dow,
        dow_avgs=json.dumps(dow_avgs),
    )
    return render(request, 'reports/forecast.html', ctx)


# ── EXPORT VIEWS ─────────────────────────────────────────────────
@login_required
def export_pdf(request):
    shop = get_current_shop(request)
    start, end, period = get_date_range(request)
    report = request.GET.get('report', 'overview')
    from .utils import generate_report_pdf
    pdf_bytes = generate_report_pdf(shop, start, end, report)
    response = HttpResponse(pdf_bytes, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="proto_report_{report}_{start}_{end}.pdf"'
    return response


@login_required
def export_csv(request):
    import csv
    shop = get_current_shop(request)
    start, end, period = get_date_range(request)
    report = request.GET.get('report', 'overview')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="proto_{report}_{start}_{end}.csv"'
    writer = csv.writer(response)

    if report == 'sales':
        writer.writerow(['Sale Number', 'Date', 'Time', 'Customer', 'Payment', 'Subtotal', 'Tax', 'Total', 'Cashier'])
        for s in Sale.objects.filter(shop=shop, created_at__date__range=(start, end), status='completed').select_related('customer', 'cashier').order_by('-created_at'):
            writer.writerow([
                s.sale_number,
                s.created_at.strftime('%Y-%m-%d'),
                s.created_at.strftime('%H:%M'),
                s.customer.name if s.customer else 'Walk-in',
                s.get_payment_method_display(),
                s.subtotal, s.tax_amount, s.total,
                s.cashier.get_full_name() or s.cashier.username,
            ])
    elif report == 'products':
        writer.writerow(['Product', 'Category', 'UOM', 'Qty Sold', 'Revenue', 'COGS', 'Profit', 'Margin %'])
        sales_qs = Sale.objects.filter(shop=shop, created_at__date__range=(start, end), status='completed')
        rows = SaleItem.objects.filter(sale__in=sales_qs).values(
            'product__name', 'product__category__name', 'product__uom__short_name'
        ).annotate(qty=Sum('quantity'), revenue=Sum('line_total'), cogs=Sum(F('buying_price') * F('quantity'))).order_by('-revenue')
        for r in rows:
            profit = float(r['revenue'] or 0) - float(r['cogs'] or 0)
            margin = round(profit / float(r['revenue']) * 100, 1) if r['revenue'] else 0
            writer.writerow([r['product__name'], r['product__category__name'], r['product__uom__short_name'], r['qty'], r['revenue'], r['cogs'], round(profit, 2), margin])
    elif report == 'expenses':
        writer.writerow(['Date', 'Category', 'Amount', 'Description'])
        for e in Expense.objects.filter(shop=shop, date__range=(start, end)).select_related('category').order_by('-date'):
            writer.writerow([e.date, e.category.name if e.category else '-', e.amount, e.description])
    elif report == 'stock':
        writer.writerow(['Product', 'Category', 'UOM', 'Qty', 'Buying Price', 'Cost Value', 'Retail Value', 'Status'])
        for l in StockLevel.objects.filter(shop=shop).select_related('product', 'product__category', 'product__uom'):
            status = 'Out of stock' if l.quantity <= 0 else ('Low' if l.is_low else 'OK')
            writer.writerow([
                l.product.name, l.product.category.name if l.product.category else '-',
                l.product.uom.short_name if l.product.uom else '-',
                l.quantity, l.product.buying_price,
                l.quantity * l.product.buying_price,
                l.quantity * l.product.selling_price,
                status,
            ])
    else:
        writer.writerow(['Metric', 'Value'])
        sales_qs = Sale.objects.filter(shop=shop, created_at__date__range=(start, end), status='completed')
        revenue  = sales_qs.aggregate(t=Sum('total'))['t'] or 0
        expenses = Expense.objects.filter(shop=shop, date__range=(start, end)).aggregate(t=Sum('amount'))['t'] or 0
        writer.writerows([
            ['Period', f'{start} to {end}'],
            ['Revenue', revenue],
            ['Transactions', sales_qs.count()],
            ['Expenses', expenses],
        ])

    return response


@login_required
def export_excel(request):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse('openpyxl not installed', status=500)

    shop = get_current_shop(request)
    start, end, period = get_date_range(request)
    report = request.GET.get('report', 'overview')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = report.title()

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='6146C1')

    def write_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 18

    if report == 'sales':
        write_header(ws, ['Sale #', 'Date', 'Time', 'Customer', 'Payment', 'Subtotal', 'Tax', 'Total'])
        for row, s in enumerate(Sale.objects.filter(shop=shop, created_at__date__range=(start, end), status='completed').select_related('customer').order_by('-created_at'), 2):
            ws.append([s.sale_number, s.created_at.strftime('%Y-%m-%d'), s.created_at.strftime('%H:%M'),
                       s.customer.name if s.customer else 'Walk-in', s.get_payment_method_display(),
                       float(s.subtotal), float(s.tax_amount), float(s.total)])
    elif report == 'products':
        write_header(ws, ['Product', 'Category', 'UOM', 'Qty Sold', 'Revenue (TSh)', 'COGS (TSh)', 'Profit (TSh)', 'Margin %'])
        sales_qs = Sale.objects.filter(shop=shop, created_at__date__range=(start, end), status='completed')
        rows = SaleItem.objects.filter(sale__in=sales_qs).values(
            'product__name', 'product__category__name', 'product__uom__short_name'
        ).annotate(qty=Sum('quantity'), revenue=Sum('line_total'), cogs=Sum(F('buying_price') * F('quantity'))).order_by('-revenue')
        for r in rows:
            profit = float(r['revenue'] or 0) - float(r['cogs'] or 0)
            margin = round(profit / float(r['revenue']) * 100, 1) if r['revenue'] else 0
            ws.append([r['product__name'], r['product__category__name'], r['product__uom__short_name'], r['qty'], float(r['revenue'] or 0), float(r['cogs'] or 0), round(profit, 2), margin])
    elif report == 'expenses':
        write_header(ws, ['Date', 'Category', 'Amount (TSh)', 'Description'])
        for e in Expense.objects.filter(shop=shop, date__range=(start, end)).select_related('category').order_by('-date'):
            ws.append([str(e.date), e.category.name if e.category else '-', float(e.amount), e.description])
    elif report == 'stock':
        write_header(ws, ['Product', 'Category', 'UOM', 'Qty', 'Buy Price', 'Cost Value', 'Retail Value', 'Status'])
        for l in StockLevel.objects.filter(shop=shop).select_related('product', 'product__category', 'product__uom'):
            status = 'Out' if l.quantity <= 0 else ('Low' if l.is_low else 'OK')
            ws.append([l.product.name, l.product.category.name if l.product.category else '-',
                       l.product.uom.short_name if l.product.uom else '-',
                       l.quantity, float(l.product.buying_price),
                       float(l.quantity * l.product.buying_price),
                       float(l.quantity * l.product.selling_price), status])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(len(str(col[0].value or '')), 12) + 2

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="proto_{report}_{start}_{end}.xlsx"'
    wb.save(response)
    return response
