"""
Bulk product upload utilities for Proto v3.
Supports CSV and Excel (.xlsx) files.

Expected columns (case-insensitive):
  name, category, selling_price, buying_price,
  description, track_stock, low_stock_threshold,
  variant_type_1, variant_value_1, variant_price_1,   (optional)
  variant_type_2, variant_value_2, variant_price_2,   (optional)
  initial_stock                                        (optional)
"""

import csv
import io
from decimal import Decimal, InvalidOperation
from django.utils.text import slugify
from .models import Product, Category, ProductVariant, VariantType, VariantAttribute, ProductVariantType
from stock.models import StockLevel, StockMovement


def parse_decimal(val, default=0):
    try:
        return Decimal(str(val).replace(',', '').strip())
    except (InvalidOperation, ValueError):
        return Decimal(str(default))


def parse_bool(val, default=True):
    if isinstance(val, bool):
        return val
    return str(val).strip().lower() not in ('false', '0', 'no', 'n', '')


def get_or_create_category(name, shop):
    name = (name or '').strip()
    if not name:
        return None
    slug = slugify(name)
    base = slug
    n = 1
    while Category.objects.filter(slug=slug).exclude(shop=shop).exists():
        slug = f"{base}-{n}"
        n += 1
    cat, _ = Category.objects.get_or_create(
        name__iexact=name, shop=shop,
        defaults={'name': name, 'slug': slug, 'shop': shop}
    )
    return cat


def process_row(row, shop, user):
    """
    Process one data row dict. Returns (product, created, error_msg).
    """
    name = (row.get('name') or row.get('Name') or '').strip()
    if not name:
        return None, False, 'Missing product name'

    selling_price = parse_decimal(row.get('selling_price') or row.get('Selling Price') or 0)
    buying_price  = parse_decimal(row.get('buying_price')  or row.get('Buying Price')  or 0)

    if selling_price <= 0:
        return None, False, f'"{name}": selling price must be greater than 0'

    category_name = row.get('category') or row.get('Category') or ''
    category = get_or_create_category(category_name, shop)

    description     = (row.get('description')         or row.get('Description')         or '').strip()
    track_stock     = parse_bool(row.get('track_stock') or row.get('Track Stock'), True)
    low_threshold   = int(parse_decimal(row.get('low_stock_threshold') or row.get('Low Stock Threshold') or 10))
    initial_stock   = int(parse_decimal(row.get('initial_stock') or row.get('Initial Stock') or 0))
    uom_name        = (row.get('uom') or row.get('UOM') or row.get('Unit') or '').strip()

    # Resolve UOM
    uom = None
    if uom_name:
        from units.models import UnitOfMeasure
        uom = UnitOfMeasure.objects.filter(name__iexact=uom_name, is_active=True).first()
        if not uom:
            uom = UnitOfMeasure.objects.filter(short_name__iexact=uom_name, is_active=True).first()

    # Create or update product
    product, created = Product.objects.get_or_create(
        name__iexact=name, shop=shop,
        defaults={
            'name': name,
            'shop': shop,
            'category': category,
            'selling_price': selling_price,
            'buying_price': buying_price,
            'description': description,
            'track_stock': track_stock,
            'low_stock_threshold': low_threshold,
        }
    )
    if not created:
        # Update existing
        product.selling_price      = selling_price
        product.buying_price       = buying_price
        product.description        = description or product.description
        product.track_stock        = track_stock
        product.low_stock_threshold = low_threshold
        if uom:
            product.uom = uom
        if category:
            product.category = category
        product.save()

    # Handle variants (variant_type_1/value_1, variant_type_2/value_2 ...)
    for i in range(1, 6):
        vtype_name = (row.get(f'variant_type_{i}') or row.get(f'Variant Type {i}') or '').strip()
        vvalue     = (row.get(f'variant_value_{i}') or row.get(f'Variant Value {i}') or '').strip()
        vprice     = row.get(f'variant_price_{i}') or row.get(f'Variant Price {i}') or None
        if not vtype_name or not vvalue:
            break

        vtype, _ = VariantType.objects.get_or_create(name=vtype_name.title())
        ProductVariantType.objects.get_or_create(product=product, variant_type=vtype)

        # Check if this exact variant already exists
        existing_attr = VariantAttribute.objects.filter(
            variant__product=product, variant_type=vtype, value__iexact=vvalue
        ).first()

        if not existing_attr:
            variant = ProductVariant.objects.create(
                product=product,
                selling_price=parse_decimal(vprice) if vprice else None,
            )
            VariantAttribute.objects.create(variant=variant, variant_type=vtype, value=vvalue)
            product.has_variants = True

    product.save(update_fields=['has_variants'])

    # Set initial stock (only for new products or if column is provided)
    if initial_stock > 0 and track_stock:
        sl, _ = StockLevel.objects.get_or_create(
            product=product, variant=None, shop=shop,
            defaults={'quantity': 0}
        )
        if created or initial_stock > sl.quantity:
            old = sl.quantity
            sl.quantity = initial_stock
            sl.save()
            StockMovement.objects.create(
                product=product, shop=shop,
                movement_type='opening',
                quantity=initial_stock - old,
                quantity_before=old,
                quantity_after=initial_stock,
                reference='Bulk upload',
                created_by=user,
            )

    return product, created, None


def process_csv(file_obj, shop, user):
    """Process a CSV file object. Returns (results, errors)."""
    results = []
    errors  = []
    try:
        text = file_obj.read().decode('utf-8-sig')  # handle BOM
        reader = csv.DictReader(io.StringIO(text))
        for i, row in enumerate(reader, start=2):
            product, created, err = process_row(row, shop, user)
            if err:
                errors.append(f'Row {i}: {err}')
            else:
                results.append({'product': product, 'created': created})
    except Exception as e:
        errors.append(f'File error: {e}')
    return results, errors


def process_excel(file_obj, shop, user):
    """Process an .xlsx file object. Returns (results, errors)."""
    results = []
    errors  = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active
        headers = [str(c.value or '').strip() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_dict = {headers[j]: row[j] for j in range(len(headers))}
            product, created, err = process_row(row_dict, shop, user)
            if err:
                errors.append(f'Row {row_idx}: {err}')
            else:
                results.append({'product': product, 'created': created})
    except ImportError:
        errors.append('openpyxl is required for Excel upload. Run: pip install openpyxl')
    except Exception as e:
        errors.append(f'File error: {e}')
    return results, errors


def generate_template_csv():
    """Return a CSV template as string for users to download."""
    headers = [
        'name', 'category', 'uom', 'selling_price', 'buying_price',
        'description', 'track_stock', 'low_stock_threshold', 'initial_stock',
        'variant_type_1', 'variant_value_1', 'variant_price_1',
        'variant_type_2', 'variant_value_2', 'variant_price_2',
    ]
    example = [
        'Gucci Handbag', 'Bags', 'Piece', '85000', '65000',
        'Premium leather handbag', 'True', '10', '20',
        'Color', 'Black', '',
        'Size', 'Medium', '',
    ]
    out = io.StringIO()
    writer = csv.writer(out)
    writer.writerow(headers)
    writer.writerow(example)
    writer.writerow(['Shea Butter 250ml', 'Hair Care', 'Millilitre', '12000', '5500',
                     'Natural shea butter', 'True', '15', '50',
                     'Size', '250ml', '', '', '', ''])
    return out.getvalue()
